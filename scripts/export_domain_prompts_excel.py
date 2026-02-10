#!/usr/bin/env python3
"""Export all domain-enriched extraction prompts to Excel.

Generates the ACTUAL prompts that ExtractionService would use for each domain,
applying the same enrichment logic as get_domain_enriched_extraction_prompts().

Output: data/evaluation/results/domain_extraction_prompts.xlsx
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from src.prompts.extraction_prompts import (
    DSPY_OPTIMIZED_ENTITY_PROMPT,
    DSPY_OPTIMIZED_RELATION_PROMPT,
    get_domain_enriched_extraction_prompts,
)


def main():
    # Load seed domains
    yaml_path = Path(__file__).resolve().parent.parent / "data" / "seed_domains.yaml"
    with open(yaml_path) as f:
        catalog = yaml.safe_load(f)

    domains = catalog.get("domains", [])
    print(f"Loaded {len(domains)} domains from {yaml_path.name}")

    # Create workbook
    wb = Workbook()

    # ── Sheet 1: Overview ──────────────────────────────────────────────
    ws_overview = wb.active
    ws_overview.title = "Overview"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    wrap_align = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Overview headers
    overview_headers = [
        "Domain ID",
        "Name",
        "DDC",
        "FORD",
        "Entity Sub-Types (count)",
        "Entity Sub-Types",
        "Relation Hints (count)",
        "Relation Hints",
        "Has JSON Example (Entity)",
        "Has JSON Example (Relation)",
        "Prompt Issues",
    ]
    for col, header in enumerate(overview_headers, 1):
        cell = ws_overview.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # ── Sheet 2: Entity Prompts ────────────────────────────────────────
    ws_entity = wb.create_sheet("Entity Prompts")
    entity_headers = [
        "Domain ID",
        "Name",
        "Full Entity Prompt",
        "Char Count",
        "Injected Sub-Types Section",
    ]
    for col, header in enumerate(entity_headers, 1):
        cell = ws_entity.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # ── Sheet 3: Relation Prompts ──────────────────────────────────────
    ws_relation = wb.create_sheet("Relation Prompts")
    relation_headers = [
        "Domain ID",
        "Name",
        "Full Relation Prompt",
        "Char Count",
        "Injected Hints Section",
    ]
    for col, header in enumerate(relation_headers, 1):
        cell = ws_relation.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # ── Sheet 4: Generic (no domain) ──────────────────────────────────
    ws_generic = wb.create_sheet("Generic (no domain)")
    ws_generic.cell(row=1, column=1, value="Type").font = header_font
    ws_generic.cell(row=1, column=2, value="Full Prompt").font = header_font
    ws_generic.cell(row=1, column=3, value="Char Count").font = header_font

    ws_generic.cell(row=2, column=1, value="Entity Prompt (generic)")
    ws_generic.cell(row=2, column=2, value=DSPY_OPTIMIZED_ENTITY_PROMPT)
    ws_generic.cell(row=2, column=2).alignment = wrap_align
    ws_generic.cell(row=2, column=3, value=len(DSPY_OPTIMIZED_ENTITY_PROMPT))

    ws_generic.cell(row=3, column=1, value="Relation Prompt (generic)")
    ws_generic.cell(row=3, column=2, value=DSPY_OPTIMIZED_RELATION_PROMPT)
    ws_generic.cell(row=3, column=2).alignment = wrap_align
    ws_generic.cell(row=3, column=3, value=len(DSPY_OPTIMIZED_RELATION_PROMPT))

    ws_generic.column_dimensions["A"].width = 25
    ws_generic.column_dimensions["B"].width = 120
    ws_generic.column_dimensions["C"].width = 12

    # ── Process each domain ────────────────────────────────────────────
    issue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row_idx, domain in enumerate(domains, 2):
        domain_id = domain.get("domain_id", "unknown")
        name = domain.get("name", "")
        ddc = domain.get("ddc_code", "")
        ford = domain.get("ford_code", "")
        sub_types = domain.get("entity_sub_types", [])
        sub_type_mapping = domain.get("entity_sub_type_mapping", {})
        relation_hints = domain.get("relation_hints", [])

        # Generate enriched prompts (same logic as production)
        entity_prompt, relation_prompt = get_domain_enriched_extraction_prompts(
            domain=domain_id,
            entity_sub_types=sub_types,
            entity_sub_type_mapping=sub_type_mapping,
            relation_hints=relation_hints,
        )

        # Check for issues
        issues = []

        # Issue 1: Entity prompt has no JSON example
        entity_has_json_example = '"name":' in entity_prompt or '"name"' in entity_prompt
        if not entity_has_json_example:
            issues.append("Entity: no JSON example")

        # Issue 2: Relation prompt JSON example
        relation_has_json_example = (
            '"subject"' in relation_prompt and '"relation"' in relation_prompt
        )
        # (relation prompt always has the generic example)

        # Issue 3: No sub-types defined
        if not sub_types:
            issues.append("No entity sub-types")

        # Issue 4: No relation hints
        if not relation_hints:
            issues.append("No relation hints")

        # Issue 5: Sub-type mapping missing entries
        unmapped = [st for st in sub_types if st not in sub_type_mapping]
        if unmapped:
            issues.append(f"Unmapped sub-types: {unmapped}")

        # Issue 6: Relation hints don't use arrow format consistently
        bad_hints = [h for h in relation_hints if "→" not in h]
        if bad_hints:
            issues.append(f"Hints without → format: {bad_hints}")

        # Issue 7: Domain-specific relation types NOT in universal 22 types
        universal_relations = {
            "PART_OF",
            "CONTAINS",
            "INSTANCE_OF",
            "TYPE_OF",
            "EMPLOYS",
            "MANAGES",
            "FOUNDED_BY",
            "OWNS",
            "LOCATED_IN",
            "CAUSES",
            "ENABLES",
            "REQUIRES",
            "LEADS_TO",
            "PRECEDES",
            "FOLLOWS",
            "USES",
            "CREATES",
            "IMPLEMENTS",
            "DEPENDS_ON",
            "SIMILAR_TO",
            "ASSOCIATED_WITH",
            "RELATED_TO",
            "RELATES_TO",
        }
        hint_rel_types = []
        for h in relation_hints:
            parts = h.split("→")
            if parts:
                rel_type = parts[0].strip()
                hint_rel_types.append(rel_type)
        non_universal = [r for r in hint_rel_types if r not in universal_relations]
        if non_universal:
            issues.append(f"Non-universal rel types in hints: {non_universal}")

        # Extract the domain-specific type sections for separate columns
        injected_entity = ""
        if "Entity types for " in entity_prompt:
            try:
                start = entity_prompt.index("Entity types for ")
                end = entity_prompt.index("\n\nRules:")
                injected_entity = entity_prompt[start:end].strip()
            except ValueError:
                injected_entity = "(parse error)"

        injected_relation = ""
        if "Relation types for " in relation_prompt:
            try:
                start = relation_prompt.index("Relation types for ")
                end = relation_prompt.index("\n\nRules:")
                injected_relation = relation_prompt[start:end].strip()
            except ValueError:
                injected_relation = "(parse error)"

        # ── Write Overview row ──
        overview_row = [
            domain_id,
            name,
            ddc,
            ford,
            len(sub_types),
            ", ".join(sub_types) if sub_types else "(none)",
            len(relation_hints),
            "\n".join(relation_hints) if relation_hints else "(none)",
            "YES" if entity_has_json_example else "NO",
            "YES" if relation_has_json_example else "NO",
            "; ".join(issues) if issues else "OK",
        ]
        for col, value in enumerate(overview_row, 1):
            cell = ws_overview.cell(row=row_idx, column=col, value=value)
            cell.alignment = wrap_align
            cell.border = thin_border
            # Color code issues
            if col == 9 and value == "NO":  # Entity JSON example
                cell.fill = issue_fill
            if col == 11 and value != "OK":  # Issues column
                cell.fill = warn_fill if "Non-universal" in str(value) else issue_fill

        # ── Write Entity Prompt row ──
        ws_entity.cell(row=row_idx, column=1, value=domain_id).border = thin_border
        ws_entity.cell(row=row_idx, column=2, value=name).border = thin_border
        c = ws_entity.cell(row=row_idx, column=3, value=entity_prompt)
        c.alignment = wrap_align
        c.border = thin_border
        ws_entity.cell(row=row_idx, column=4, value=len(entity_prompt)).border = thin_border
        c2 = ws_entity.cell(
            row=row_idx,
            column=5,
            value=injected_entity if injected_entity else "(none — uses generic)",
        )
        c2.alignment = wrap_align
        c2.border = thin_border
        if not injected_entity:
            c2.fill = warn_fill

        # ── Write Relation Prompt row ──
        ws_relation.cell(row=row_idx, column=1, value=domain_id).border = thin_border
        ws_relation.cell(row=row_idx, column=2, value=name).border = thin_border
        c = ws_relation.cell(row=row_idx, column=3, value=relation_prompt)
        c.alignment = wrap_align
        c.border = thin_border
        ws_relation.cell(row=row_idx, column=4, value=len(relation_prompt)).border = thin_border
        c2 = ws_relation.cell(
            row=row_idx,
            column=5,
            value=injected_relation if injected_relation else "(none — uses generic)",
        )
        c2.alignment = wrap_align
        c2.border = thin_border
        if not injected_relation:
            c2.fill = warn_fill

    # ── Column widths ──────────────────────────────────────────────────
    ws_overview.column_dimensions["A"].width = 25
    ws_overview.column_dimensions["B"].width = 30
    ws_overview.column_dimensions["C"].width = 10
    ws_overview.column_dimensions["D"].width = 8
    ws_overview.column_dimensions["E"].width = 12
    ws_overview.column_dimensions["F"].width = 50
    ws_overview.column_dimensions["G"].width = 12
    ws_overview.column_dimensions["H"].width = 60
    ws_overview.column_dimensions["I"].width = 15
    ws_overview.column_dimensions["J"].width = 15
    ws_overview.column_dimensions["K"].width = 60

    ws_entity.column_dimensions["A"].width = 25
    ws_entity.column_dimensions["B"].width = 30
    ws_entity.column_dimensions["C"].width = 120
    ws_entity.column_dimensions["D"].width = 12
    ws_entity.column_dimensions["E"].width = 60

    ws_relation.column_dimensions["A"].width = 25
    ws_relation.column_dimensions["B"].width = 30
    ws_relation.column_dimensions["C"].width = 120
    ws_relation.column_dimensions["D"].width = 12
    ws_relation.column_dimensions["E"].width = 60

    # ── Freeze panes ───────────────────────────────────────────────────
    ws_overview.freeze_panes = "A2"
    ws_entity.freeze_panes = "A2"
    ws_relation.freeze_panes = "A2"

    # ── Save ───────────────────────────────────────────────────────────
    output_path = (
        Path(__file__).resolve().parent.parent
        / "data"
        / "evaluation"
        / "results"
        / "domain_extraction_prompts.xlsx"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    print(f"\nExcel saved to: {output_path}")
    print(f"Domains: {len(domains)}")
    print(f"Sheets: Overview, Entity Prompts, Relation Prompts, Generic (no domain)")

    # Print summary of issues
    print("\n--- Issue Summary ---")
    total_issues = 0
    for domain in domains:
        domain_id = domain.get("domain_id", "unknown")
        sub_types = domain.get("entity_sub_types", [])
        relation_hints = domain.get("relation_hints", [])
        hint_rel_types = []
        for h in relation_hints:
            parts = h.split("→")
            if parts:
                hint_rel_types.append(parts[0].strip())
        non_universal = [r for r in hint_rel_types if r not in universal_relations]
        if non_universal:
            print(f"  {domain_id}: non-universal rel types in hints: {non_universal}")
            total_issues += 1
    print(f"\nTotal domains with non-universal relation hints: {total_issues}/{len(domains)}")


if __name__ == "__main__":
    main()
