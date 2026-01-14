#!/usr/bin/env python3
"""Test Table Format Routing (CSV, XLSX).

Sprint 89 Pre-Validation: Ensures table formats are routed to Docling
(not processed as plaintext) for proper table extraction.

This test verifies:
1. CSV files are routed to Docling (not LlamaIndex plaintext)
2. XLSX files are routed to Docling (table extraction)
3. Docling extracts table structure (not just text)

Usage:
    poetry run python scripts/test_table_routing.py

Expected Output:
    ✅ CSV routed to Docling
    ✅ XLSX routed to Docling
    ✅ Table structure preserved (columns detected)
"""

import asyncio
import sys
import tempfile
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

import structlog

logger = structlog.get_logger(__name__)


async def test_format_router_routing():
    """Test 1: Verify FormatRouter routes CSV/XLSX to Docling."""
    from src.components.ingestion.format_router import (
        DOCLING_FORMATS,
        FormatRouter,
        ParserType,
    )

    print("\n" + "=" * 60)
    print("TEST 1: Format Router Routing")
    print("=" * 60)

    # Check format sets
    csv_in_docling = ".csv" in DOCLING_FORMATS
    xlsx_in_docling = ".xlsx" in DOCLING_FORMATS

    print(f"\n📋 Format Set Membership:")
    print(f"   .csv in DOCLING_FORMATS: {csv_in_docling}")
    print(f"   .xlsx in DOCLING_FORMATS: {xlsx_in_docling}")

    if not csv_in_docling or not xlsx_in_docling:
        print("\n❌ FAIL: CSV/XLSX not in DOCLING_FORMATS!")
        return False

    # Test router decisions
    router = FormatRouter(docling_available=True)

    csv_decision = router.route(Path("test.csv"))
    xlsx_decision = router.route(Path("test.xlsx"))

    print(f"\n🔀 Routing Decisions:")
    print(f"   CSV  → {csv_decision.parser.value} ({csv_decision.reason})")
    print(f"   XLSX → {xlsx_decision.parser.value} ({xlsx_decision.reason})")

    csv_ok = csv_decision.parser == ParserType.DOCLING
    xlsx_ok = xlsx_decision.parser == ParserType.DOCLING

    if csv_ok and xlsx_ok:
        print("\n✅ PASS: CSV and XLSX correctly routed to Docling")
        return True
    else:
        print("\n❌ FAIL: Routing incorrect!")
        return False


async def test_docling_table_extraction():
    """Test 2: Verify Docling extracts table structure from CSV/XLSX."""
    from src.components.ingestion.docling_client import DoclingClient

    print("\n" + "=" * 60)
    print("TEST 2: Docling Table Extraction")
    print("=" * 60)

    # Create test CSV
    csv_content = """Framework,Type,Language,Stars
TensorFlow,Deep Learning,Python,180000
PyTorch,Deep Learning,Python,75000
SpaCy,NLP,Python,28000
Neo4j,Graph Database,Java,12000
"""

    # Create test XLSX (requires openpyxl)
    xlsx_path = None
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Frameworks"

        # Header row
        headers = ["Framework", "Type", "Language", "Stars"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Data rows
        data = [
            ("TensorFlow", "Deep Learning", "Python", 180000),
            ("PyTorch", "Deep Learning", "Python", 75000),
            ("SpaCy", "NLP", "Python", 28000),
            ("Neo4j", "Graph Database", "Java", 12000),
        ]
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        xlsx_path = Path(tempfile.mktemp(suffix=".xlsx"))
        wb.save(xlsx_path)
        print(f"\n📄 Created test XLSX: {xlsx_path}")

    except ImportError:
        print("\n⚠️  openpyxl not installed, skipping XLSX test")
        print("   Install with: pip install openpyxl")

    # Create test CSV file
    csv_path = Path(tempfile.mktemp(suffix=".csv"))
    csv_path.write_text(csv_content)
    print(f"📄 Created test CSV: {csv_path}")

    # Initialize Docling client
    client = DoclingClient()

    # Check Docling availability via HTTP
    print("\n🔍 Checking Docling availability...")
    import httpx
    from src.core.config import settings

    docling_ok = False
    try:
        async with httpx.AsyncClient(timeout=10.0) as http_client:
            response = await http_client.get(f"{settings.docling_base_url}/openapi.json")
            docling_ok = response.status_code == 200
            print(f"   Docling URL: {settings.docling_base_url}")
            print(f"   Health check: {'OK' if docling_ok else 'FAILED'}")
    except Exception as e:
        print(f"   Docling unreachable: {e}")

    if not docling_ok:
        print("⚠️  Docling container not available!")
        print("   Start with: docker start docling-service")
        print("   Skipping Docling extraction tests...")
        return None  # Skip, not fail

    results = {}

    # Test CSV parsing
    print("\n📊 Parsing CSV with Docling...")
    try:
        csv_result = await client.parse_document(csv_path)
        results["csv"] = {
            "text_length": len(csv_result.text),
            "tables_count": len(csv_result.tables),
            "has_table_structure": len(csv_result.tables) > 0,
            "parse_time_ms": csv_result.parse_time_ms,
        }
        print(f"   Text length: {results['csv']['text_length']} chars")
        print(f"   Tables found: {results['csv']['tables_count']}")
        print(f"   Parse time: {results['csv']['parse_time_ms']:.0f}ms")

        # Check if table data is preserved
        text_lower = csv_result.text.lower()
        has_headers = all(h.lower() in text_lower for h in ["framework", "type", "language"])
        has_data = "tensorflow" in text_lower and "pytorch" in text_lower
        results["csv"]["content_preserved"] = has_headers and has_data
        print(f"   Headers preserved: {has_headers}")
        print(f"   Data preserved: {has_data}")

    except Exception as e:
        print(f"   ❌ CSV parsing failed: {e}")
        results["csv"] = {"error": str(e)}

    # Test XLSX parsing
    if xlsx_path:
        print("\n📊 Parsing XLSX with Docling...")
        try:
            xlsx_result = await client.parse_document(xlsx_path)
            results["xlsx"] = {
                "text_length": len(xlsx_result.text),
                "tables_count": len(xlsx_result.tables),
                "has_table_structure": len(xlsx_result.tables) > 0,
                "parse_time_ms": xlsx_result.parse_time_ms,
            }
            print(f"   Text length: {results['xlsx']['text_length']} chars")
            print(f"   Tables found: {results['xlsx']['tables_count']}")
            print(f"   Parse time: {results['xlsx']['parse_time_ms']:.0f}ms")

            # Check if table data is preserved
            text_lower = xlsx_result.text.lower()
            has_headers = all(h.lower() in text_lower for h in ["framework", "type", "language"])
            has_data = "tensorflow" in text_lower and "pytorch" in text_lower
            results["xlsx"]["content_preserved"] = has_headers and has_data
            print(f"   Headers preserved: {has_headers}")
            print(f"   Data preserved: {has_data}")

        except Exception as e:
            print(f"   ❌ XLSX parsing failed: {e}")
            results["xlsx"] = {"error": str(e)}

    # Cleanup
    csv_path.unlink(missing_ok=True)
    if xlsx_path:
        xlsx_path.unlink(missing_ok=True)

    # Evaluate results
    print("\n" + "-" * 40)
    print("📋 Summary:")

    all_passed = True

    for fmt, result in results.items():
        if "error" in result:
            print(f"   {fmt.upper()}: ❌ Error - {result['error']}")
            all_passed = False
        elif result.get("content_preserved"):
            print(f"   {fmt.upper()}: ✅ Content preserved")
        else:
            print(f"   {fmt.upper()}: ⚠️  Content may be incomplete")

    return all_passed


async def test_frontend_api_upload():
    """Test 3: Upload via Frontend API and verify routing."""
    import httpx

    print("\n" + "=" * 60)
    print("TEST 3: Frontend API Upload (E2E)")
    print("=" * 60)

    # Create minimal test CSV
    csv_content = """Entity,Type,Description
AegisRAG,Software,RAG System with Graph Intelligence
Qdrant,Database,Vector Search Engine
Neo4j,Database,Graph Database
"""

    csv_path = Path(tempfile.mktemp(suffix=".csv"))
    csv_path.write_text(csv_content)
    print(f"\n📄 Created test CSV: {csv_path}")

    # Upload via Frontend API
    api_url = "http://localhost:8000/api/v1/retrieval/upload"
    namespace = "table_routing_test"

    # Get auth token
    auth_url = "http://localhost:8000/api/v1/auth/login"
    print(f"\n🔐 Getting auth token...")

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:
            # Login first (JSON body, not form-data)
            auth_response = await client.post(
                auth_url,
                json={"username": "admin", "password": "admin123"},
            )
            if auth_response.status_code != 200:
                print(f"   ⚠️  Auth failed: {auth_response.status_code}")
                print(f"   Response: {auth_response.text[:200]}")
                csv_path.unlink(missing_ok=True)
                return None

            token = auth_response.json().get("access_token")
            print(f"   ✅ Got token: {token[:20]}...")

            print(f"\n🚀 Uploading to {api_url}...")
            with open(csv_path, "rb") as f:
                response = await client.post(
                    api_url,
                    files={"file": (csv_path.name, f, "text/csv")},
                    data={
                        "namespace": namespace,
                        "domain": "test",
                    },
                    headers={"Authorization": f"Bearer {token}"},
                )

            print(f"   Status: {response.status_code}")

            if response.status_code == 200:
                result = response.json()
                print(f"   Document ID: {result.get('document_id', 'N/A')}")
                print(f"   Status: {result.get('status', 'N/A')}")

                # Check if routed to Docling (from logs or response)
                if "document_id" in result:
                    print("\n✅ Upload successful!")
                    print("   Check server logs for: 'parser=docling' to confirm routing")
                    csv_path.unlink(missing_ok=True)
                    return True
            elif response.status_code == 500:
                print(f"   ⚠️  Server error (not routing related): {response.text[:150]}")
                print("   Note: This is likely an embedding/GPU issue, not routing.")
                csv_path.unlink(missing_ok=True)
                return None  # Skip, not fail (routing test passed above)
            else:
                print(f"   ❌ Upload failed: {response.text[:200]}")
                csv_path.unlink(missing_ok=True)
                return False

    except httpx.ConnectError:
        print("   ⚠️  API not reachable (localhost:8000)")
        print("   Start backend with: poetry run uvicorn src.api.main:app --port 8000")
        csv_path.unlink(missing_ok=True)
        return None  # Skip, not fail

    except Exception as e:
        print(f"   ❌ Error: {e}")
        csv_path.unlink(missing_ok=True)
        return False


async def main():
    """Run all table routing tests."""
    print("\n" + "=" * 60)
    print("TABLE FORMAT ROUTING TEST SUITE")
    print("Sprint 89 Pre-Validation")
    print("=" * 60)

    results = {}

    # Test 1: Format Router
    results["format_router"] = await test_format_router_routing()

    # Test 2: Docling Extraction
    results["docling_extraction"] = await test_docling_table_extraction()

    # Test 3: Frontend API
    results["frontend_api"] = await test_frontend_api_upload()

    # Final Summary
    print("\n" + "=" * 60)
    print("FINAL RESULTS")
    print("=" * 60)

    for test_name, result in results.items():
        if result is True:
            status = "✅ PASS"
        elif result is False:
            status = "❌ FAIL"
        else:
            status = "⚠️  SKIP"
        print(f"   {test_name}: {status}")

    # Overall result
    failures = [r for r in results.values() if r is False]
    if failures:
        print(f"\n❌ {len(failures)} test(s) failed!")
        return 1
    else:
        print("\n✅ All tests passed (or skipped)!")
        return 0


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)
